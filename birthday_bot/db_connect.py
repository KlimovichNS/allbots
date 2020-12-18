import sqlite3
from sqlite3 import Error
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re

def sql_connection():
    try:
        con = sqlite3.connect('bdaylist.db')
        return con
    except Error:
        print(Error)

def get_users():
    conn = sql_connection()
    conn.row_factory = lambda cursor, row: row[0]
    c = conn.cursor()
    ids = c.execute('SELECT id FROM userlist').fetchall()
    conn.close()
    
    return ids


def users_table(con):
    cursorObj = con.cursor()
    cursorObj.execute("CREATE TABLE IF NOT EXISTS userlist(id integer PRIMARY KEY, name text);")
    con.commit()
    con.close()
    
def sql_insert(con, entities):
    # entities = (uid, name)
    cursorObj = con.cursor()
    cursorObj.execute('INSERT OR REPLACE INTO userlist(id, name) VALUES(?, ?)', entities)
    con.commit()
    con.close()


def slugify(text, lower=1):
    if lower == 1:
        if not text is None:
            text = str(text).strip().lower()
    text = re.sub(r'[^\w _-]+', '', str(text))
    
    text = re.sub(r'[- ]+', '_', str(text))
    
    return text

def new_table(uid):
    con = sql_connection()
    bd_list=r'docs/' + str(uid)+'.xlsx'
    wb = load_workbook(filename=bd_list)

    sheets = wb.sheetnames

    for sheet in sheets:
        ws = wb[sheet] 
        
        columns= []
        query = 'CREATE TABLE IF NOT EXISTS list' + str(uid) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'
        for row in next(ws.rows):
            query += ', ' + slugify(row.value) + ' TEXT'
            print(slugify(row.value))
            columns.append(slugify(row.value))
        query += ');'
        
        con.execute(query)

        tup = []
        for i, rows in enumerate(ws):
            tuprow = []
            if i == 0:
                continue
            for row in rows:
                tuprow.append(str(row.value).strip()) if str(row.value).strip() != None else tuprow.append('')
            tup.append(tuple(tuprow))


        insQuery1 = 'INSERT INTO list' + str(uid) + '('
        insQuery2 = ''
        for col in columns:
            insQuery1 += col + ', '
            insQuery2 += '?, '
        insQuery1 = insQuery1[:-2] + ') VALUES('
        insQuery2 = insQuery2[:-2] + ')'
        insQuery = insQuery1 + insQuery2

        con.executemany(insQuery, tup)
        #rename_query = 'ALTER TABLE ' +'лист1' + ' RENAME TO ' + '"' + str(uid) + '"'
        #con.execute(rename_query)
        con.commit()

    con.close()

